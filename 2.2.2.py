import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import re
from prettytable import PrettyTable
import datetime as DT


class Vacancy:
    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.rubles_currency[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

    rubles_currency = {
        "RUR": 1,
        "EUR": 59.90,
        "KZT": 0.13,
        "BYR": 23.91,
        "AZN": 35.68,
        "GEL": 21.74,
        "KGS": 0.76,
        "USD": 60.66,
        "UZS": 0.0055,
        "UAH": 1.64
    }


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            headlines = next(reader)
            for line in reader:
                if len(headlines) == len(line) and '' not in line:
                    yield dict(zip(headlines, line))

    def get_statistic(self):
        vacancies_count = 0
        salary = {}
        salary_vacancy = {}
        salary_place = {}

        for vacancy_dict in self.csv_reader():
            vacancy = Vacancy(vacancy_dict)
            self.augmentation(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.augmentation(salary_vacancy, vacancy.year, [vacancy.salary_average])
            self.augmentation(salary_place, vacancy.area_name, [vacancy.salary_average])
            vacancies_count += 1

        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        vacancies_number_by_name = dict([(key, len(value)) for key, value in salary_vacancy.items()])

        if not salary_vacancy:
            salary_vacancy = dict([(key, [0]) for key, value in salary.items()])
            vacancies_number_by_name = dict([(key, 0) for key, value in vacancies_number.items()])

        statistics, statistics2, statistics_p, statistics_v = self.statistics_sorting(salary, salary_place, salary_vacancy, vacancies_count)
        return statistics, vacancies_number, statistics_v, vacancies_number_by_name, statistics_p, statistics2

    def statistics_sorting(self, salary, salary_place, salary_vacancy, vacancies_count):
        statistics = self.mean(salary)
        statistics_p = self.mean(salary_place)
        statistics_v = self.mean(salary_vacancy)
        statistics1 = {}
        for year, salaries in salary_place.items():
            statistics1[year] = round(len(salaries) / vacancies_count, 4)
        statistics1 = list(filter(lambda x: x[-1] >= 0.01, [(key, value) for key, value in statistics1.items()]))
        statistics1.sort(key=lambda x: x[-1], reverse=True)
        statistics2 = statistics1.copy()
        statistics1 = dict(statistics1)
        statistics_p = list(
            filter(lambda x: x[0] in list(statistics1.keys()), [(key, value) for key, value in statistics_p.items()]))
        statistics_p.sort(key=lambda x: x[-1], reverse=True)
        statistics_p = dict(statistics_p[:10])
        statistics2 = dict(statistics2[:10])
        return statistics, statistics2, statistics_p, statistics_v

    @staticmethod
    def mean(dict):
        dict_new = {}
        for key, values in dict.items():
            dict_new[key] = int(sum(values) / len(values))
        return dict_new

    @staticmethod
    def augmentation(dictionary, key, sum):
        dictionary[key] = dictionary[key] + sum if key in dictionary else sum

    @staticmethod
    def print_result(stat_salary, stat_vac, stat_salary_prof, stat_vac_prof, stat_salary_place, stat_salary_place_prof):
        print('Динамика уровня зарплат по годам: {0}'.format(stat_salary))
        print('Динамика количества вакансий по годам: {0}'.format(stat_vac))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stat_salary_prof))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stat_vac_prof))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stat_salary_place))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stat_salary_place_prof))


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stat_salary, stat_vac, stat_salary_prof, stat_vac_prof, stat_salary_place, stat_salary_place_prof = dataset.get_statistic()
        dataset.print_result(stat_salary, stat_vac, stat_salary_prof, stat_vac_prof, stat_salary_place, stat_salary_place_prof)
        report = File(self.vacancy_name, stat_salary, stat_vac, stat_salary_prof, stat_vac_prof, stat_salary_place, stat_salary_place_prof)
        report.generate_excel()


class File:
    def __init__(self, vacancy_name, stat_salary, stat_vac, stat_salary_prof, stat_vac_prof, stat_salary_place, stat_salary_place_prof):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stat_salary
        self.stats2 = stat_vac
        self.stats3 = stat_salary_prof
        self.stats4 = stat_vac_prof
        self.stats5 = stat_salary_place
        self.stats6 = stat_salary_place_prof

    def generate_excel(self):
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        salary_av = 'Средняя зарплата - ' + self.vacancy_name
        count_vac ='Количество вакансий - ' + self.vacancy_name
        ws1.append(['Год', 'Средняя зарплата', salary_av, 'Количество вакансий', count_vac])
        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]
        for i, column_width in enumerate(column_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2
        self.parameters_excel(data, ws1, ws2)
        self.wb.save('report.xlsx')

    def parameters_excel(self, data, ws1, ws2):
        font_bold = Font(bold=True)
        for k in 'ABCDE':
            ws1[k + '1'].font = font_bold
            ws2[k + '1'].font = font_bold
        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'
        thin = Side(border_style='thin', color='00000000')
        for line in range(len(data)):
            for k in 'ABDE':
                ws2[k + str(line + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)
        self.stats1[1] = 1
        for line, _ in enumerate(self.stats1):
            for k in 'ABCDE':
                ws1[k + str(line + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)


class InputConnect:
    def __init__(self):
        self.params = InputConnect.get_params()


def csv_readers(fileName):
    file_csv = csv.reader(open(fileName, encoding='utf_8_sig'))
    try:
        list_data = [x for x in file_csv]
        titles = list_data[0]
        values = [x for x in list_data[1:] if x.count('') == 0 and len(x) == len(titles)]
        return titles, values
    except:
        print('Пустой файл')
        exit()


def formatter(row):
    for vacation in row:
        for key, value in vacation.items():
            if vacation[key] == 'False':
                vacation[key] = 'Нет'
            elif vacation[key] == 'True':
                vacation[key] = 'Да'
            try:
                vacation[key] = dictionary[value]
            except:
                pass

        datetime = DT.datetime.strptime(vacation['published_at'][0:10].replace('-', ''), '%Y%m%d').date()
        vacation['published_at'] = datetime.strftime('%d.%m.%Y')
        if vacation['salary_gross'] == 'Да':
            vacation['salary_gross'] = 'Без вычета налогов'
        else:
            vacation['salary_gross'] = 'С вычетом налогов'
        vacation['salary_from'] = f"{'{0:,}'.format(int(float(vacation['salary_from']))).replace(',', ' ')} - " \
                                 f"{'{0:,}'.format(int(float(vacation['salary_to']))).replace(',', ' ')} " \
                                 f"({vacation['salary_currency']}) ({vacation['salary_gross']})"
        del vacation['salary_currency'], vacation['salary_gross'], vacation['salary_to']
    return row


def print_vacancies(data_vacancies, lines, inputFields):
    if len(data_vacancies) == 0:
        print('Нет данных')
        return
    table = PrettyTable()
    table.field_names = ['Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия', 'Компания', 'Оклад', 'Название региона', 'Дата публикации вакансии']
    table.max_width = 20
    table.align = 'l'
    table.hrules = 1
    for item in formatter(data_vacancies):
        column = []
        for key, value in item.items():
            if len(value) > 100:
                column.append(value[:100] + '...')
            else:
                column.append(value)
        table.add_row(column)
    table.add_autoindex('№')
    if inputFields != '':
        fieldList = inputFields.split(', ')
        fieldList.append('№')
        if lines.isdigit() == True:
            print(table.get_string(start = int(lines) - 1, fields = fieldList))
        elif len(lines.split()) == 2:
            borders = lines.split()
            print(table.get_string(start = int(borders[0]) - 1, end = int(borders[1]) - 1, fields = fieldList))
        elif lines == '':
            print(table.get_string(fields = fieldList))
    else:
        if lines.isdigit() == True:
            print(table.get_string(start=int(lines) - 1))
        elif len(lines.split()) == 2:
            borders = lines.split()
            print(table.get_string(start=int(borders[0]) - 1, end=int(borders[1]) - 1))
        elif lines == '':
            print(table)


def csv_filter(reader, list_naming):
    vacList = []
    for i in range(0, len(reader)):
        for j in range(0, len(reader[i])):
            reader[i][j] = clean_string(reader[i][j])
    for value in reader:
        vacationDictionary = {}
        count = 0
        for field in value:
            vacationDictionary[list_naming[count]] = field
            count += 1
        vacList.append(vacationDictionary)
    return vacList


def clean_string(text):
    text = re.sub(r'<[^>]*>', '', text).replace('\r\n', ' ').strip()
    return re.sub(' +', ' ', text)


dictionary = {
    'name': 'Название',
    'description': 'Описание',
    'key_skills': 'Навыки',
    'experience_id': 'Опыт работы',
    'premium': 'Премиум-вакансия',
    'employer_name': 'Компания',
    'salary_from': 'Оклад',
    'salary_to': 'Верхняя граница вилки оклада',
    'salary_gross': 'Оклад указан до вычета налогов',
    'salary_currency': 'Идентификатор валюты оклада',
    'area_name': 'Название региона',
    'published_at': 'Дата публикации вакансии',
    "noExperience": "Нет опыта",
    "between1And3": "От 1 года до 3 лет",
    "between3And6": "От 3 до 6 лет",
    "moreThan6": "Более 6 лет",
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
    "UZS": "Узбекский сум"
}

a = input()
if a == 'Вакансии' and __name__ == '__main__':
    print('Введите данные для печати: ')
    titles, values = csv_readers(input())
    print_vacancies(csv_filter(values, titles), input(), input())
if a == 'Статистика' and __name__ == '__main__':
    InputConnect()

