import csv, re, datetime, os

from prettytable import PrettyTable, ALL
from xlsx2html import xlsx2html

import matplotlib
import matplotlib.pyplot as plt
import numpy as np

from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.styles import Font, Side, Border, Alignment
from openpyxl.workbook import Workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


class InputParam:
    """Получение вводимых параметров
    Attributes:
        param(Tuple[str, str, str, str, List[int], str]): вводимые параметры
    """
    def __init__(self):
        """Инициализирует объект InputParam
       """
        self.param = InputParam.get_params()

    @staticmethod
    def get_params():
        """Получает вводимые параметры
       Returns:
           param(Tuple[str, str, str, str, List[int], str]): вводимые параметры
       """
        file_name = input('Введите название файла: ')
        parameter = input('Введите параметр фильтрации: ')
        sort_param = input('Введите параметр сортировки: ')
        sort_reverse = input('Обратный порядок сортировки (Да / Нет): ')
        gap_rows = list(map(int, input('Введите диапазон вывода: ').split()))
        new_columns = input('Введите требуемые столбцы: ')
        if os.stat(file_name).st_size == 0:
            print("Пустой файл")
            exit()
        return file_name, parameter, sort_param, sort_reverse, gap_rows, new_columns

class InputConnectVacancy:
    """Формирование таблицы PrettyTable с вакансиями
    Attributes:
        filter_p(str): Параметр фильтрации
        sort_p(str): Параметр сортировки
        sort_r(str or bool): Параметр обратной сортировки
        gap_rows (List[int]): Промежуток выводимых колонок
        columns(str or List[str]): Названия выводимых колонок
    """
    def __init__(self, filter_p, sort_p, sort_r, gap_rows, columns):
        """Инициализирует объект InputConnectVacancy
        Args:
            filter_p(str): Параметр фильтрации
            sort_p(str): Параметр сортировки
            sort_r(str or bool): Параметр обратной сортировки
            gap_rows (List[int]): Промежуток выводимых колонок
            columns(str or List[str]): Названия выводимых колонок
        """
        self.filter_p = filter_p
        self.sort_p = sort_p
        self.sort_r = sort_r
        self.gap_rows = gap_rows
        self.columns = columns

    def formatter(self, vacancy):
        """Осуществляет форматирование необходимых полей
        Args:
            vacancy (Vacancy): Вакансия
        Returns:
            List[Any]: Список отформатированных полей вакансии
        """
        def check_salary(salary):
            """Форматирует зараплату
            Args:
                salary(Salary): Информация о зарплате
            Returns:
                str: Отформатированная информация о зарплате
            """
            new_salary_from = int(float(salary.salary_from))
            new_salary_to = int(float(salary.salary_to))
            new_salary_from = f'{new_salary_from // 1000} {str(new_salary_from)[-3:]}' if float(
                salary.salary_from) >= 1000 else new_salary_from
            new_salary_to = f'{new_salary_to // 1000} {str(new_salary_to)[-3:]}' if float(
                salary.salary_to) >= 1000 else new_salary_to
            new_salary_gross = "Без вычета налогов" if transl_bool[salary.salary_gross] == 'Да' else "С вычетом налогов"
            result = f'{new_salary_from} - {new_salary_to} ({transl_currency[salary.salary_currency]}) ({new_salary_gross})'
            return result


        def check_published(date):
            """Форматирует дату публикации
            Args:
                date(str): Дата публикации
            Returns:
                str: Отформатированная дата публикации
            """
            return datetime.datetime.strptime(date, '%Y-%m-%dT%H:%M:%S%z').strftime('%d.%m.%Y')

        result = [vacancy.name, vacancy.description, "\n".join(vacancy.key_skills),
                  transl_experience[vacancy.experience_id],
                  transl_bool[vacancy.premium], vacancy.employer_name, check_salary(vacancy.salary), vacancy.area_name,
                  check_published(vacancy.published_at)]
        return result

    def check_param(self):
        """Проверка параметров на корректность
        """
        if ': ' not in self.filter_p and self.filter_p != '':
            print('Формат ввода некорректен')
            exit()
        self.filter_p = self.filter_p.split(': ')
        if len(self.filter_p) == 2 and self.filter_p[0] not in list(dict_translation.values()):
            print('Параметр поиска некорректен')
            exit()
        if self.sort_p not in list(dict_translation.values()) and self.sort_p != '':
            print('Параметр сортировки некорректен')
            exit()
        if self.sort_r not in ['Да', 'Нет', '']:
            print('Порядок сортировки задан некорректно')
            exit()
        self.sort_r = self.sort_r == 'Да'
        if len(self.columns) != 0:
            self.columns = self.columns.split(', ')
            self.columns.insert(0, "№")

    def data_filter(self, list_vacancies, f_param):
        """Осуществляет фильтрацию списка вакансий
        Args:
            list_vacancies(List[Vacancy]): Список вакансий
            f_param(List[str]): Параметры фильтрации
        Returns:
            List[Vacancy]: Отфильтрованный список вакансий
        """
        if f_param[0] == 'Навыки':
            f_param[1] = f_param[1].split(', ')
        if f_param[0] == 'Оклад':
            list_vacancies = list(
                filter(lambda vac: int(vac.salary.salary_to) >= int(f_param[1]) >= int(vac.salary.salary_from),
                       list_vacancies))
        elif f_param[0] == 'Навыки':
            list_vacancies = list(
                filter(lambda vac: all(skill in vac.key_skills for skill in f_param[1]), list_vacancies))
        elif f_param[0] == 'Опыт работы':
            list_vacancies = list(
                filter(lambda vac: f_param[1] == transl_experience[vac.experience_id], list_vacancies))
        elif f_param[0] == 'Идентификатор валюты оклада':
            list_vacancies = list(
                filter(lambda vac: f_param[1] == transl_currency[vac.salary.salary_currency], list_vacancies))
        elif f_param[0] == 'Премиум-вакансия':
            list_vacancies = list(filter(lambda vac: f_param[1] == transl_bool[vac.premium], list_vacancies))
        elif f_param[0] == 'Дата публикации вакансии':
            list_vacancies = list(
                filter(lambda vac: f_param[1] == datetime.datetime.strptime(vac.published_at, '%Y-%m-%dT%H:%M:%S%z')
                       .strftime('%d.%m.%Y'), list_vacancies))
        else:
            list_vacancies = list(
                filter(lambda vac: f_param[1] == vac.__getattribute__(reverse_transl[f_param[0]]), list_vacancies))
        return list_vacancies

    def sort_data(self, list_vacancies, sort_par, sort_reverse):
        """Осуществляет сортировку списка вакансий по параметрам
        Args:
            list_vacancies(List[Vacancy]): Список вакансий
            sort_par(str): Параметр сортировки
            sort_reverse(bool): Параметр обратной сортировки
        Returns:
            List[Vacancy]: Отсортированный список вакансий
        """
        if sort_par == 'Навыки':
            list_vacancies.sort(key=lambda vac: len(vac.key_skills), reverse=sort_reverse)
        elif sort_par == 'Оклад':
            list_vacancies.sort(
                key=lambda vac: vac.salary.do_rub((float(vac.salary.salary_from)) + float(vac.salary.salary_to)) / 2,
                reverse=sort_reverse)
        elif sort_par == 'Дата публикации вакансии':
            list_vacancies.sort(key=lambda vac: datetime.datetime.strptime(vac.published_at, '%Y-%m-%dT%H:%M:%S%z'),
                                reverse=sort_reverse)
        elif sort_par == 'Опыт работы':
            list_vacancies.sort(key=lambda vac: weight_experience[vac.experience_id], reverse=sort_reverse)
        else:
            list_vacancies.sort(key=lambda vac: vac.__getattribute__(reverse_transl[sort_par]), reverse=sort_reverse)
        return list_vacancies

    def print_vacancies(self, list_vac):
        """Формирует таблицу PrettyTable по отсортированному и отфильтрованному списку вакансий
        Args:
            list_vac(List[Vacancy]): Список вакансий
        """
        self.gap_rows.append(len(list_vac) + 1)
        list_vac = list_vac if len(self.filter_p) != 2 else self.data_filter(list_vac, self.filter_p)
        list_vac = 'Ничего не найдено' if len(list_vac) == 0 else list_vac
        if type(list_vac) is str:
            print(list_vac)
            return
        list_vac = list_vac if len(self.sort_p) == 0 else self.sort_data(list_vac, self.sort_p, self.sort_r)
        new_columns = list(reverse_transl.keys())[:-1]
        new_columns.insert(0, "№")
        vacancies_table = PrettyTable(new_columns)
        vacancies_table.hrules = ALL
        for i in range(len(list_vac)):
            new_dict = self.formatter(list_vac[i])
            new_dict = list(map(lambda i: f'{i[:100]}...' if len(i) > 100 else i, new_dict))
            new_dict.insert(0, i + 1)
            vacancies_table.add_row(new_dict)
        vacancies_table.align = 'l'
        vacancies_table.max_width = 20
        if len(self.columns) >= 2 and len(self.gap_rows) > 1:
            vacancies_table = vacancies_table.get_string(start=self.gap_rows[0] - 1, end=self.gap_rows[1] - 1, fields=self.columns)
        elif len(self.gap_rows) > 1:
            vacancies_table = vacancies_table.get_string(start=self.gap_rows[0] - 1, end=self.gap_rows[1] - 1)
        elif len(self.columns) >= 2:
            vacancies_table = vacancies_table.get_string(fields=self.columns)
        print(vacancies_table)


class Report:
    """Формирует отчеты и графики
    """
    def generate_excel(self, key_vac, statistics):
        """Формирует Excel-таблицу
        Args:
            key_vac(str): Название вакансии
            statistics(List[Dict[str, str]]): Статистика по вакансиям
        """
        wb = Workbook()
        thins = Side(border_style="thin", color="000000")
        sh1 = wb['Sheet']
        sh1.title = 'Статистика по годам'
        wb.create_sheet('Статистика по городам')
        col_names1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {key_vac}', 'Количество вакансий', f'Количество вакансий - {key_vac}']
        for i, column in enumerate(col_names1):
            sh1.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        for year, value in statistics[0].items():
            sh1.append([year, value, statistics[1][year], statistics[2][year], statistics[3][year]])
        for column in sh1.columns:
            length = max(len(str(cell.value)) for cell in column)
            sh1.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        sh2 = wb['Статистика по городам']
        col_names2 = ['Город', 'Уровень зарплат', '  ', 'Город', 'Доля вакансий']
        for i, column in enumerate(col_names2):
            sh2.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        activeStat = list(statistics[4].items())
        for i in range(10):
            activeStat[i] += tuple(statistics[5].items())[i]
        for city1, value1, city2, value2 in activeStat:
            sh2.append([city1, value1, '  ', city2, value2])
        for i in range(2, 12):
            sh2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00
        for column in sh2.columns:
            length = max(len(str(cell.value)) for cell in column)
            sh2.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                cell.alignment = Alignment(horizontal='center')
                if cell.value != '  ':
                    cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        wb.save('report.xlsx')

    def generate_image(self, key_vac, statistics):
        """Формирует PNG-файл с графиками
        Args:
            key_vac(str): Название вакансии
            statistics(List[Dict[str, str]]): Статистика по вакансиям
        """
        matplotlib.rc('font', size=8)
        labels = statistics[0].keys()
        avg_salary = statistics[0].values()
        avg_salary_vac = statistics[1].values()
        count_vacancies = statistics[2].values()
        vacancies = statistics[3].values()
        cities = list(statistics[4].keys())
        salary_city = statistics[4].values()
        vac_city_names = list(statistics[5].values())
        vac_city_names = [1-sum(vac_city_names)] + vac_city_names
        vac_city =  ['Другие'] + list(statistics[5].keys())

        for i in range(len(cities)):
            cities[i] = cities[i].replace(' ', '\n')
            if '-' in cities[i]:
                cities[i] = '-\n'.join(cities[i].split('-'))

        x = np.arange(len(labels))
        width = 0.35

        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        ax1.bar(x - width / 2, avg_salary, width, label='средняя з/п')
        ax1.bar(x + width / 2, avg_salary_vac, width, label=f'з/п {key_vac}')
        ax1.set_title('Уровень зарплат по годам')
        ax1.set_xticks(x, labels, fontsize=8, rotation=90)
        ax1.legend(loc="upper left", fontsize=8)
        ax1.grid(axis='y')

        ax2.bar(x - width / 2, count_vacancies, width, label='Количество вакансий')
        ax2.bar(x + width / 2, vacancies, width, label=f'Количество вакансий {key_vac}')
        ax2.set_title('Количество вакансий по годам')
        ax2.set_xticks(x, labels, fontsize=8, rotation=90)
        ax2.legend(loc="upper left", fontsize=8)
        ax2.grid(axis='y')

        y_pos = np.arange(len(cities))
        ax3.barh(y_pos, salary_city, align='center')
        ax3.set_yticks(y_pos, labels=cities, fontsize=6)
        ax3.invert_yaxis()
        ax3.set_title('Уровень зарплат по городам')
        ax3.grid(axis='x')

        ax4.pie(vac_city_names, labels=vac_city, radius=1, textprops={"fontsize": 6})
        ax4.set_title('Доля вакансий по городам')
        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self, name_vac):
        """Формирует PDF-файл с графиками и отчетами
        Args:
            name_vac(str): Название вакансии
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        out = xlsx2html('report.xlsx', sheet='Статистика по годам')
        out.seek(0)
        code = out.read()
        out1 = xlsx2html('report.xlsx', sheet='Статистика по городам')
        out1.seek(0)
        code1 = out1.read()

        pdf_template = template.render({'name_vac': name_vac, 'table1': code, 'table2': code1})

        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltoox\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class InputConectStatistics:
    """Формирует статистику по вакансиям
    """
    def get_salary_level(self, list_vac, key, param_vac=''):
        """Формирует статистику, связанную с зарплатами
        Args:
            list_vac(List[Vacancy]): Список вакансий
            key(str): Поле вакансии
            param_vac(str): Параметр с названием вакансии
        Returns:
            Dict[str, str]: Cтатистика, связанная с зарплатами
        """
        result = {}
        for vacancy in list_vac:
            if vacancy.__getattribute__(key) not in result.keys():
                result[vacancy.__getattribute__(key)] = []
        list_vac = list((filter(lambda vac: param_vac in vac.name, list_vac))) if param_vac != '' else list_vac
        for vac in list_vac:
            result[vac.__getattribute__(key)].append(vac.salary.do_rub(float(vac.salary.salary_from) + float(vac.salary.salary_to)) / 2)
        for key in result.keys():
            result[key] = 0 if len(result[key]) == 0 else int(sum(result[key]) // len(result[key]))
        return result

    def get_count_vacancy(self, list_vac, key, dataSet, param_vac=''):
        """Формирует статистику, связанную с количеством вакансий
        Args:
            list_vac(List[Vacancy]): Список вакансий
            key(str): Поле вакансии
            dataSet(DataSet): Данные из файла
            param_vac(str): Параметр с названием вакансии
        Returns:
            Dict[str, str]: Cтатистика, связанная с количеством вакансий
        """
        result = {}
        for vacancy in list_vac:
            if vacancy.__getattribute__(key) not in result.keys():
                result[vacancy.__getattribute__(key)] = 0
        list_vac = list((filter(lambda vac: param_vac in vac.name, list_vac))) if param_vac != '' else list_vac
        for vac in list_vac:
            result[vac.__getattribute__(key)] += 1
        if key == 'area_name':
            for key in result.keys():
                result[key] = round(result[key] / len(dataSet.vacancies_objects), 4)
        return result
    def check_published(self, date):
        """Форматирует дату публикации
        Args:
            date(List[str]): Дата публикации
        Returns:
            str: Отформатированная дата публикации
        """
        return datetime.datetime.strptime(date, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')

    def print_res(self, dataSet, outputer, vacancy_name):
        """Осуществляет вывод статистики по вакансиям
        Args:
            dataSet(DataSet): Данные из файла
            outputer(InputConectStatistics): Статистика по вакансиям
            vacancy_name: Название вакансии
        Returns:
            List[Any]: Список статистик
        """
        for vac in dataSet.vacancies_objects:
            vac.published_at = int(self.check_published(vac.published_at))
        res_year = dict(
            sorted(outputer.get_salary_level(dataSet.vacancies_objects, "published_at").items(), key=lambda x: x[0]))
        print(f'Динамика уровня зарплат по годам: {res_year}')
        res_vac = dict(
            sorted(outputer.get_count_vacancy(dataSet.vacancies_objects, "published_at", dataSet).items(), key=lambda x: x[0]))
        print(f'Динамика количества вакансий по годам: {res_vac}')
        res_year_vac = dict(sorted(outputer.get_salary_level(dataSet.vacancies_objects, "published_at", vacancy_name).items(),
                          key=lambda x: x[0]))
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {res_year_vac}')
        res_vac_count = dict(sorted(outputer.get_count_vacancy(dataSet.vacancies_objects, "published_at", dataSet, vacancy_name).items(),
                          key=lambda x: x[0]))
        print(f'Динамика количества вакансий по годам для выбранной профессии: {res_vac_count}')
        dict_city_count = {}
        for vac in dataSet.vacancies_objects:
            if vac.area_name not in dict_city_count.keys():
                dict_city_count[vac.area_name] = 0
            dict_city_count[vac.area_name] += 1
        new_vacancies_objects = list(
            filter(lambda vac: int(len(dataSet.vacancies_objects) * 0.01) <= dict_city_count[vac.area_name],
                   dataSet.vacancies_objects))
        res_city = dict(sorted(outputer.get_salary_level(new_vacancies_objects, "area_name").items(), key=lambda x: x[1],
                          reverse=True)[:10])
        print(f'Уровень зарплат по городам (в порядке убывания): {res_city}')
        res_city_vac = dict(sorted(outputer.get_count_vacancy(new_vacancies_objects, "area_name", dataSet).items(), key=lambda x: x[1],
                          reverse=True)[:10])
        print(f'Доля вакансий по городам (в порядке убывания): {res_city_vac}')
        return [res_year, res_year_vac, res_vac, res_vac_count, res_city, res_city_vac]

class DataSet:
    """Формирует список вакансий
    Attributes:
        file_name(str): Имя файла
        vacancies_objects(List[Vacancy]): Список вакансий
    """
    def __init__(self, file_name):
        """Инициализирует объект DataSet
        Args:
            file_name(str): Имя файла
        """
        self.file_name = file_name
        self.vacancies_objects = [Vacancy(vac) for vac in self.csv_filer(*self.csv_reader(file_name))]

    def csv_filer(self, reader, list_naming):
        """Формирует список словарей по вакансиям
        Args:
            reader(List[List[str]]): Данные из файла
            list_naming: Поля вакансии
        Returns:
            List[Dict[str, str]]: Список словарей по вакансиям
        """
        res = [self.formatting_str(x) for x in reader if len(x) == len(list_naming) and x.count('') == 0]
        dic_result = []
        for line in res:
            dic = {}
            for i in range(len(line)):
                dic[list_naming[i]] = line[i]
            dic_result.append(dic)
        return dic_result

    def csv_reader(self, file_name):
        """Считывает данные из файла
        Args:
            file_name(str): Название файла
        Returns:
            Tuple[List[str], List[List[str]]]: Названия полей и соответствующие им данные
        """
        file_csv = open(file_name, encoding='utf_8_sig')
        reader_csv = csv.reader(file_csv)
        list_data = []
        for str in reader_csv:
            list_data.append(str)
        columns = list_data[0]
        data = list_data[1:]
        return data, columns

    def formatting_str(self, element):
        """Очищает строку от HTML кода
        Args:
            element(List(str)): Список строк, которые необходимо очистить
        Returns:
            str: Список очищенных строк
        """
        for i in range(len(element)):
            element[i] = re.sub(r'<.*?>', '', element[i])
            if '\n' not in element[i]:
                element[i] = " ".join(element[i].split())
        return element


class Salary:
    """Класс для представления зарплаты.
    Attributes:
        salary_from(str): Нижняя граница вилки оклада
        salary_to(str): Верхняя граница вилки оклада
        salary_gross(str): Наличие налога
        salary_currency(str): Валюта оклада
    """
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """Инициализирует объект Salary
        Args:
            salary_from(str): Нижняя граница вилки оклада
            salary_to(str): Верхняя граница вилки оклада
            salary_gross(str): Наличие налога
            salary_currency(str): Валюта оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def do_rub(self, salary):
        """Переводит в рубли при помощи словаря - currency_to_rub.
        Args:
            salary(float): Зарплата в другой валюте
        Returns:
            float: Перевод в рубли
        >>> Salary('10', '1000', 'true', 'EUR').do_rub(1000.0)
        59900.0
        >>> Salary('10', '1000', 'true', 'RUR').do_rub(1000)
        1000.0
        >>> Salary('10', '1000', 'true', '123').do_rub(1000.0)
        Traceback (most recent call last):
            ...
        KeyError: '123'
        """
        return salary * float(currency_to_rub[self.salary_currency])

class Vacancy:
    """Класс для представления вакансии
    Attributes:
        name (str): Название вакансии
        description (str): Описание вакансии
        key_skills (List[str]): Ключевые навыки
        experience_id(str): Требуемый опыт
        premium(str): Премиум-вакансия
        employer_name(str): Название компании
        salary(Salary): Информация о зарплате
        area_name(str): Название города
        published_at(str): Дата публикации
    """
    def __init__(self, data_vac):
        """Инициализирует объект Vacancy, проверяет наличие полей
        Args:
             data_vac (Dict[str, str]): Словарь, в котором описывается профессия. Ключи - названия полей, значения - соответствующая информация по полю
        """
        self.name = data_vac['name']
        self.description = "Нет данных" if 'description' not in data_vac.keys() else data_vac['description']
        self.key_skills = "Нет данных" if 'key_skills' not in data_vac.keys() else data_vac['key_skills'].split('\n')
        self.experience_id = "Нет данных" if 'experience_id' not in data_vac.keys() else data_vac['experience_id']
        self.premium = "Нет данных" if 'premium' not in data_vac.keys() else data_vac['premium']
        self.employer_name = "Нет данных" if 'employer_name' not in data_vac.keys() else data_vac['employer_name']
        salary_gross = "Нет данных" if 'salary_gross' not in data_vac.keys() else data_vac['salary_gross']
        self.salary = Salary(data_vac['salary_from'], data_vac['salary_to'], salary_gross, data_vac['salary_currency'])
        self.area_name = data_vac['area_name']
        self.published_at = data_vac['published_at']


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

dict_translation = {"name": "Название",
                    "description": "Описание",
                    "key_skills": "Навыки",
                    "experience_id": "Опыт работы",
                    "premium": "Премиум-вакансия",
                    "employer_name": "Компания",
                    "salary_from": "Нижняя граница вилки оклада",
                    "salary_to": "Верхняя граница вилки оклада",
                    "salary_gross": "Оклад указан до вычета налогов",
                    "salary_currency": "Идентификатор валюты оклада",
                    "area_name": "Название региона",
                    "published_at": "Дата публикации вакансии",
                    "Оклад": "Оклад"}

transl_experience = {"noExperience": "Нет опыта",
                     "between1And3": "От 1 года до 3 лет",
                     "between3And6": "От 3 до 6 лет",
                     "moreThan6": "Более 6 лет",
                     "Нет данных": "Нет данных"}

weight_experience = {
    "noExperience": 0,
    "between1And3": 1,
    "between3And6": 2,
    "moreThan6": 3}

transl_currency = {"AZN": "Манаты",
                  "BYR": "Белорусские рубли",
                  "EUR": "Евро",
                  "GEL": "Грузинский лари",
                  "KGS": "Киргизский сом",
                  "KZT": "Тенге",
                  "RUR": "Рубли",
                  "UAH": "Гривны",
                  "USD": "Доллары",
                  "UZS": "Узбекский сум"}

transl_bool = {"True": "Да",
               "TRUE": "Да",
               "False": "Нет",
               "FALSE": "Нет",
               "Нет данных": "Нет данных"}

reverse_transl = {"Название": "name",
                  "Описание": "description",
                  "Навыки": "key_skills",
                  "Опыт работы": "experience_id",
                  "Премиум-вакансия": "premium",
                  "Компания": "employer_name",
                  "Оклад": "Оклад",
                  "Название региона": "area_name",
                  "Дата публикации вакансии": "published_at",
                  "Идентификатор валюты оклада": "salary_currency"}

if __name__ == '__main__':
    type_output = input('Введите данные для печати: ')
    if type_output == "Статистика":
        file_name = input('Введите название файла: ')
        vacancy_name = input('Введите название профессии: ')
        if os.stat(file_name).st_size == 0:
            print("Пустой файл")
            exit()
        outputer = InputConectStatistics()
        dataSet = DataSet(file_name)
        if len(dataSet.vacancies_objects) == 0:
            print('Нет данных')
            exit()
        statistics = outputer.print_res(dataSet, outputer, vacancy_name)
        rp = Report()
        rp.generate_excel(vacancy_name, statistics)
        rp.generate_image(vacancy_name, statistics)
        rp.generate_pdf(vacancy_name)
    if type_output == "Вакансии":
        params = InputParam()
        outputer = InputConnectVacancy(params.param[1], params.param[2], params.param[3], params.param[4], params.param[5])
        outputer.check_param()
        if params.param is not None:
            dataSet = DataSet(params.param[0])
            if len(dataSet.vacancies_objects) == 0:
                print('Нет данных')
                exit()
            outputer.print_vacancies(dataSet.vacancies_objects)